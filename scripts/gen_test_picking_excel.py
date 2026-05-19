# -*- coding: utf-8 -*-
"""
gen_test_picking_excel.py
-------------------------
가상(테스트용) Picking List Excel 생성 — PDF(test_picking_list)와 동일 내용.

요건:
  - 현재 RESERVED 68개 LOT 전부 대상
  - 각 LOT RESERVED 톤백의 절반((R+1)//2)을 피킹 → 본품(MT) 행
  - LOT마다 샘플(KG 1.0) 행 1줄
  - 고객 = PT LBM

형식: features/parsers/picking_excel_parser.parse_picking_list_excel() 규격
  상단 키-값 + 데이터 헤더(Lot No / Quantity / Unit / Storage location) + 데이터 행
"""
import os
import sys
import sqlite3

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
DB_PATH = os.path.join(ROOT, 'data', 'db', 'sqm_inventory.db')
OUT_PATH = os.path.join(ROOT, 'test_picking_list_20260518.xlsx')

TONBAG_MT = 0.5
STORAGE = '1001 GY logistics'


def reserved_counts():
    db = sqlite3.connect(DB_PATH)
    rows = db.execute(
        "SELECT lot_no, COUNT(*) FROM inventory_tonbag "
        "WHERE status='RESERVED' AND COALESCE(is_sample,0)=0 "
        "GROUP BY lot_no ORDER BY lot_no"
    ).fetchall()
    db.close()
    return rows


def main():
    lots = reserved_counts()
    if not lots:
        print('[FAIL] RESERVED 톤백이 없습니다.')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PickingList'
    bold = Font(bold=True)
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='C00000')
    sample_fill = PatternFill('solid', fgColor='FFF2CC')
    center = Alignment(horizontal='center')

    ws['A1'] = 'PICKING LIST'
    ws['A1'].font = Font(bold=True, size=12)
    kv = [
        ('Outbound ID', '80100501'),
        ('Sales Order', '80007501'),
        ('Customer reference', 'PTLBM-PICK-2606'),
        ('Customer', 'PT LBM'),
        ('Creation Date', '2026-05-18'),
        ('Plan Loading Date', '2026-05-25'),
    ]
    for i, (k, v) in enumerate(kv, start=2):
        ws[f'A{i}'] = k
        ws[f'A{i}'].font = bold
        ws[f'B{i}'] = v

    # 데이터 헤더 (9행)
    hdr_row = 9
    for ci, name in enumerate(['Lot No', 'Quantity', 'Unit', 'Storage location'], start=1):
        c = ws.cell(row=hdr_row, column=ci, value=name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    row = hdr_row + 1
    total_picked_tb = 0
    total_mt = 0.0
    for lot_no, r in lots:
        picked = (r + 1) // 2
        mt = round(picked * TONBAG_MT, 2)
        total_picked_tb += picked
        total_mt += mt
        # 본품(MT) 행
        ws.cell(row=row, column=1, value=lot_no)
        ws.cell(row=row, column=2, value=mt)
        ws.cell(row=row, column=3, value='MT')
        ws.cell(row=row, column=4, value=STORAGE)
        row += 1
        # 샘플(KG) 행
        for ci, v in enumerate([lot_no, 1.0, 'KG', STORAGE], start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.fill = sample_fill
        row += 1

    for ci, w in enumerate([16, 12, 8, 22], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(OUT_PATH)

    print('=== Picking List Excel 생성 완료 ===')
    print(f'  파일       : {OUT_PATH}')
    print(f'  LOT 수     : {len(lots)}개 (RESERVED 전부)')
    print(f'  본품 행    : {len(lots)}행  /  샘플 행 : {len(lots)}행  /  총 {row-hdr_row-1}행')
    print(f'  피킹 톤백  : {total_picked_tb}개  /  본품 합계 : {round(total_mt,2)} MT')
    print(f'  Picking No : PTLBM-PICK-2606  /  고객: PT LBM')


if __name__ == '__main__':
    main()
