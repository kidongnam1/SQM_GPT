# -*- coding: utf-8 -*-
"""위치재고조회 엑셀 ↔ 랙 구조 매핑 검토 (1회성 분석)."""
import sys, io, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from engine_modules.warehouse_cell_logic import validate_cell_location, CELL_RE

P = r'D:\program\SQM_inventory\위치재고조회_20260518_140004.xlsx'
wb = openpyxl.load_workbook(P, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
wb.close()

header = rows[0]
data = rows[1:]
print('총 데이터 행:', len(data))

LOC_RE = re.compile(r'^(G\d-\d\d-\d\d-\d\d)\s*(?:\[(\d+)\])?$')
cell_owner = {}     # location -> [lot, ...]
all_loc = []
bad_fmt = []
bad_valid = []
qty_mismatch = []
bracket_vals = {}
levels_used = {}    # (동,칸,열) -> set(층)

for ri, row in enumerate(data, start=1):
    loc_cell = row[2]
    lot = row[4]
    qty = row[7]
    if loc_cell is None:
        continue
    lines = [l.strip() for l in str(loc_cell).split('\n') if l.strip()]
    sum_bracket = 0
    for ln in lines:
        m = LOC_RE.match(ln)
        if not m:
            bad_fmt.append((ri, lot, ln))
            continue
        loc = m.group(1)
        br = int(m.group(2)) if m.group(2) else None
        sum_bracket += (br or 0)
        bracket_vals[br] = bracket_vals.get(br, 0) + 1
        all_loc.append(loc)
        cell_owner.setdefault(loc, []).append(str(lot))
        v = validate_cell_location(loc)
        if not v['ok']:
            bad_valid.append((ri, lot, loc, v['reason']))
        else:
            key = (v['dong'], v['rack'], v['col'])
            levels_used.setdefault(key, set()).add(v['level'])
    # qty 정합성
    try:
        q = int(qty)
    except (TypeError, ValueError):
        q = None
    if q is not None and sum_bracket and q != sum_bracket:
        qty_mismatch.append((ri, lot, q, sum_bracket))

print()
print('=== 1. 위치 형식 ===')
print('  파싱된 위치 셀:', len(all_loc), '개 / 고유:', len(set(all_loc)))
print('  형식오류(정규식 불일치):', len(bad_fmt))
for x in bad_fmt[:10]:
    print('   행%d LOT%s ::' % (x[0], x[1]), repr(x[2]))

print()
print('=== 2. 랙구조 검증 (validate_cell_location) ===')
print('  검증실패:', len(bad_valid))
for x in bad_valid[:20]:
    print('   행%d LOT%s loc=%s :: %s' % x)

print()
print('=== 3. [N] 괄호값 분포 ===')
for k in sorted(bracket_vals, key=lambda v: (v is None, v)):
    print('  [%s] : %d 개 셀' % (k, bracket_vals[k]))

print()
print('=== 4. qty 컬럼 vs 괄호합 불일치 ===')
print('  불일치 행:', len(qty_mismatch))
for x in qty_mismatch[:15]:
    print('   행%d LOT%s : qty=%s, 괄호합=%s' % x)

print()
print('=== 5. 위치 중복 점유 (한 셀 = 2개 이상 LOT) ===')
dup = {k: v for k, v in cell_owner.items() if len(set(v)) > 1}
print('  중복 셀:', len(dup))
for k, v in list(dup.items())[:15]:
    print('   %s :: LOT %s' % (k, sorted(set(v))))

print()
print('=== 6. LOT 적재 패턴 (동/칸/열당 층 사용) ===')
lv_counts = {}
for key, lvset in levels_used.items():
    n = len(lvset)
    lv_counts[n] = lv_counts.get(n, 0) + 1
    mn, mx = min(lvset), max(lvset)
    if sorted(lvset) != list(range(mn, mx + 1)):
        print('   ⚠ 층 불연속 G%d-%02d-%02d :: %s' % (key[0], key[1], key[2], sorted(lvset)))
for n in sorted(lv_counts):
    print('  층 %d개 사용 컬럼: %d 개' % (n, lv_counts[n]))

print()
print('=== 7. 사용 범위 ===')
oks = [validate_cell_location(x) for x in set(all_loc)]
oks = [o for o in oks if o['ok']]
if oks:
    print('  동:', sorted({o['dong'] for o in oks}))
    print('  칸:', sorted({o['rack'] for o in oks}))
    print('  열:', sorted({o['col'] for o in oks}))
    print('  층:', sorted({o['level'] for o in oks}))
