# -*- coding: utf-8 -*-
"""
gen_test_woo_allocation.py
--------------------------
가상(테스트용) Allocation Excel 생성 — Woo 양식.

요건:
  - 현재 AVAILABLE 재고 68개 LOT 전부 대상
  - LOT당 4~6 톤백(1톤백=0.5MT) 배분 → 전체 ~340MT 중 약 절반(~170MT)
  - LOT 통째로 빼지 않음 (정규 10톤백 중 4~6개만)
  - LOT마다 샘플 행 1줄(QTY 0.001MT≈1kg, 파서가 is_sample 자동 인식)
  - SOLD TO = PT LBM 단일

Woo 양식 구조 (woo_202606.json 기준):
  1행 타이틀 / 2~5행 헤더블록 / 6행 컬럼 / 7행~ 데이터
  컬럼: Product · SAP NO · Date in stock · QTY (MT) · Lot No · WH ·
        Customs · Export · SOLD TO · SALE REF · Balance · GW · Remark
"""
import os
import sqlite3
import random
import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
DB_PATH = os.path.join(ROOT, 'data', 'db', 'sqm_inventory.db')
OUT_PATH = os.path.join(ROOT, 'test_woo_allocation_20260518.xlsx')

TONBAG_MT = 0.5            # 정규 톤백 1개 = 500kg = 0.5MT
SAMPLE_MT = 0.001          # 샘플 = 1kg = 0.001MT (< 0.01 → 파서 is_sample 자동)
CUSTOMER = 'PT LBM'
SALE_REF = 'PTLBM-AL-2606'
COLUMNS = ['Product', 'SAP NO', 'Date in stock', 'QTY (MT)', 'Lot No', 'WH',
           'Customs', 'Export', 'SOLD TO', 'SALE REF', 'Balance', 'GW', 'Remark']

random.seed(20260518)      # 재현성


def load_available_lots():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    rows = db.execute(
        "SELECT lot_no, sap_no, product, current_weight, stock_date, warehouse, customs "
        "FROM inventory WHERE status='AVAILABLE' ORDER BY lot_no"
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def main():
    lots = load_available_lots()
    if not lots:
        print('[FAIL] AVAILABLE 재고가 없습니다.')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Allocation'

    bold = Font(bold=True)
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    sample_fill = PatternFill('solid', fgColor='FFF2CC')
    center = Alignment(horizontal='center')

    # ── 1행: 타이틀 (파서가 customer 'PT LBM' 추출) ──
    ws['A1'] = f'Allocation - {CUSTOMER} - May 2026 / CIF Busan - Woo 양식 테스트 (LITHIUM CARBONATE)'
    ws['A1'].font = Font(bold=True, size=12)

    # ── 2~5행: 헤더 블록 ──
    ws['A2'] = 'Customer';            ws['B2'] = CUSTOMER
    ws['A3'] = 'Product';             ws['B3'] = 'LITHIUM CARBONATE'
    ws['A4'] = 'Period';              ws['B4'] = '2026-05'
    ws['A5'] = 'Note';                ws['B5'] = '테스트용 가상 배분 — LOT당 4~6톤백 + 샘플 1행'
    for r in (2, 3, 4, 5):
        ws[f'A{r}'].font = bold

    # ── 6행: 컬럼 헤더 ──
    for ci, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=6, column=ci, value=name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    # ── 7행~: 데이터 (LOT별 본품 1행 + 샘플 1행) ──
    row = 7
    total_main_mt = 0.0
    total_tonbags = 0
    sample_count = 0
    for lot in lots:
        tonbags = random.choice([4, 5, 6])
        qty_mt = round(tonbags * TONBAG_MT, 3)
        total_main_mt += qty_mt
        total_tonbags += tonbags
        balance_mt = round((lot['current_weight'] or 0) / 1000.0, 3)
        stock_date = lot['stock_date'] or '2026-05-18'

        # 본품 행
        vals = ['LITHIUM CARBONATE', lot['sap_no'], stock_date, qty_mt,
                lot['lot_no'], lot['warehouse'] or 'GY', lot['customs'] or '',
                '일반수출', CUSTOMER, SALE_REF, balance_mt,
                round(qty_mt * 1000, 1), '']
        for ci, v in enumerate(vals, start=1):
            ws.cell(row=row, column=ci, value=v)
        row += 1

        # 샘플 행 (QTY 0.001MT → 파서 is_sample 자동)
        svals = ['LITHIUM CARBONATE', lot['sap_no'], stock_date, SAMPLE_MT,
                 lot['lot_no'], lot['warehouse'] or 'GY', lot['customs'] or '',
                 '샘플', CUSTOMER, SALE_REF, balance_mt, 1.0, 'SAMPLE']
        for ci, v in enumerate(svals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.fill = sample_fill
        row += 1
        sample_count += 1

    # 컬럼 폭
    widths = [20, 13, 13, 10, 13, 6, 11, 10, 12, 16, 10, 9, 10]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(OUT_PATH)

    print('=== Woo 양식 테스트 Allocation 생성 완료 ===')
    print(f'  파일      : {OUT_PATH}')
    print(f'  LOT 수    : {len(lots)}개 (AVAILABLE 전부)')
    print(f'  본품 행   : {len(lots)}행  /  샘플 행 : {sample_count}행  /  총 {row-7}행')
    print(f'  배분 톤백 : {total_tonbags}개 (LOT당 4~6, 1톤백 0.5MT)')
    print(f'  본품 합계 : {round(total_main_mt,3)} MT  (전체 재고 ~340MT의 약 '
          f'{round(total_main_mt/340.068*100,1)}%)')
    print(f'  샘플 합계 : {round(sample_count*SAMPLE_MT,3)} MT')
    print(f'  고객사    : {CUSTOMER}  /  SALE REF : {SALE_REF}')


if __name__ == '__main__':
    main()
